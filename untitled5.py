import re
import random
import itertools
from tkinter import *
from openpyxl import *


def make_code_list(list_code=[]):
	temp=[]
	for i in range(1,len(l)):
		temp.append(l[i][0])
	temp=set(temp)
	temp=sorted(list(temp))
	for i in range(len(temp)):
		list_code.append(temp[i])

def sel_teachers_dict():
	global selected_code
	global sel_teachers
	for i in selected_code:
		l1=[]
		for j in l:
			if(j[0]==i):
				if j[2] not in l1:
					l1.append(j[2])
			sel_teachers[i]=l1


def checkInvalidComb(l=[],selected_code=[]):
    n=len(selected_code)-1
    k=0
    while(n>-1):
        i=selected_code[n]
        l3=list(i.values())[0]
        n1=len(l3)
        n3=0
        for j in l3:
            if(j==l[n]):
                k=1
            else:
                n3+=1
        if(n3==n1):
            return(0)
        n=n-1
    return(k)




def List_dict_of_Code_slot(l=[],selected_code=[]):
    for i in selected_code:
        d1={}
        l1=[]
        for j in l:
            if(j[0]==i):
                if j[1] not in l1:
                    l1.append(j[1])
                if j[1] not in dict_totalslots:
                    dict_totalslots.append(j[1])
        d1[i]=l1
        dict_sc.append(d1)
        
    

       


def List_dict_of_slot_teacher(dict_sc=[],dict_st=[],l=[]):
    n=0
    for i in selected_code:
        d={}
        q=[]
        a=dict_sc[n]
        l3=sorted(list(a.values())[0])
        for k in l3:
            l1=[]
            for j in l:
                if(i==j[0]):
                    if(j[1]==k):
                        l1.append(j[2])
            d[k]=l1
        q.append(d)
        dict_st.append(q)
        n+=1


def make_comb():
	global n
	perms = itertools.permutations(dict_totalslots,n)
	for j in perms:
		x=checkInvalidComb(j,dict_sc)
		if(x==1):
			comb.append(j)

			print(j)


#after pereferences comb stored in comb2

def return_slot_of_teacher(p,m,list_dict2=[]):
    list5=[]
    i=list_dict2[p]
    for j in i:
        l=list(j.items())
        for k in l:
            l2=k[1]
            for p in l2:
                if(p==m):
                    list5.append(k)
    return(list5)
            
                        

def index(code,selected_code=[]):
        for i in range(len(selected_code)):
                if(code==selected_code[i]):
                    return(i)



def pref_comb(p_teachers=[],dict_st=[]):
    global comb
    global comb2
    global selected_code
    comb3=[]
    for j in p_teachers.keys():
        temp=[]
        p=index(j,selected_code)
        teacher=p_teachers[j]
        list_comp=return_slot_of_teacher(p,teacher,dict_st)
        for j in list_comp:
            w=j[0]
            temp.append(w)
        for i in comb:
            flag=0
            for  j in temp:
                if(i[p]==j):
                    flag=1
            if(flag==0):
                comb3.append(i)
                
    comb3=list(set(comb3)) 
    for i in comb:
        if i not in comb3:
            comb2.append(i)






l=[]           	 	#TABLE
list_code=[]   	 	#store all the couse codes

dict_sc=[] 			#for slot :::  make dict of the slot and code as key   
dict_st=[] 			#for slot and teacher 
dict_tv=[] 			#for teacher and venue
dict_totalslots=[]  #for the total no of slots
comb=[] 	        #to store the  all combinations

global sel_teachers    #all selected subject-teachers
sel_teachers = {}

global p_teachers      # preferred subject-teachers
p_teachers = {}

global selected_code
selected_code=[] 	#to store the selected subject

global comb2
comb2=[]
# adds=[] subject list boxes  




    
def excel0():
	wb = load_workbook(filename=n1Box.get())
	ws = wb.worksheets[0]
	rlen = ws.max_row
	clen = ws.max_column
	table=[[(ws.cell(row=j+1, column=i+1)).value for i in range(clen)] for j in range(3,rlen)]
	global l
	l=table
	t=[[l[i][j] for j in [0,11,13]]for i in range(len(l))]
	l=t
	#l=[['ite1004','b1','ram'],['ite1002','b1','sham'],['ite1004','a1','guju'],['ite1004','b2','chirayu'] ,['ite1004','b2','chirayu2222'],['ite1001','b2','guru'],['ite1001','c2','singh'],['ite1005','k1','aa']]
	make_code_list(list_code)
	return(table[0])
	




#Update available subject list (after reading excel file)
def excel():
	global adds
	excel0()
	for i in list_code:
		for j in range(len(adds)):
			adds[j].insert(END,i)
	return(l[0])





#Add new subject column
def add1():
	global adds
	adds.append(Listbox(top,height=1))
	global fcp
	fcp+=1
	
	adds[-1].grid(row=3,column=fcp)
	adds[-1].insert(END, "Select Subject")
	fcp+=1
	
	for i in list_code:
			adds[-1].insert(END,i)
	
	but2.grid(row=3, column=fcp)
	fcp+=1
	but3.grid(row=3, column=fcp)
	
	
	

	
	
#Add new selected sub-teacher row
def addst():	
	global lbst
	lbst=[]
	lbst.append(Listbox(top,height=1))
	global frp
	frp+=1
	lbst[0].grid(row=frp,column=2,sticky=N)
		
	but_ttupdate = Button(text="Update", command=ttupdate).grid(row=frp, column=3)	
	
		
	lbst.append(Listbox(top,height=1))
	lbst[1].grid(row=frp,column=4)
	lbst[1].insert(1, "Select Teacher")
	
	frp+=1
	but_ttgen.grid(row=frp, column=2)
	frp+=1
	resultLabel.grid(row=frp, column=1)
	
	#Update subject-preference list
	global selected_code
	
	for i in selected_code:
		lbst[0].insert(END,i)
			
	lbtt.append(lbst)	
	return lbst


	
	
	
#Update preference: subject-teacher list box
def sel():
	global adds
	global selected_code
	selected_code.clear()
	for i in range(len(adds)):
		selected_code.insert(0,adds[i].get('active'))
		
	print(mvar.get())
	
	
	#if(mvar.get()==1):
	global lbtt
	lbtt[0][0].delete(0,10)
	selected_code=set(selected_code)
	selected_code=sorted(list(selected_code))
	
	#Update subject-preference list
	for i in selected_code:
		lbtt[0][0].insert(END,i)
	
	
	
	
	
	
	
#Update  preference: teacher list

def ttupdate():
	global lbtt
	global selected_code
	global p_teachers
	sel_teachers_dict()
	global lbst
	i=lbst[0].get('active')

	for z in lbtt:
		l1=[]
		i=z[0].get('active')
		for j in l:
			if(j[0]==i):
				if j[2] not in l1:
					l1.append(j[2])
		p_teachers[i]=l1
		
		z[1].delete(0,10)
		for j in l1:
			z[1].insert(END,j)
	
	
	
#GENERATE TIMETABLE
		
		
def generate():
	global p_teachers
	p_teachers={}
	print("Final selected teachers list")
	for t in lbtt:
		i=t[0].get('active')
		p_teachers[i]=t[1].get('active')
	for i in p_teachers.keys():
		print(i)
		print(p_teachers[i])

	print("\n\n")
	print("Subjects : ")
	print(selected_code)
	print("\n")
	print("Slots")
	
	global n	
	n=len(selected_code)
	List_dict_of_Code_slot(l,selected_code)
	List_dict_of_slot_teacher(dict_sc,dict_st,l)
	make_comb()	
	pref_comb(p_teachers,dict_st)
	
	print("\nAccording to preferred teachers\n")
	print("Subjects : ")
	print(selected_code)
	print("\n")
	print("Slots")	
	for i in comb2:
		print(i)

	


#	//////////////////////////////////////GUI				GUI				GUI//////////////////////////////////////

top = Tk()

mvar=IntVar()

resultStr= StringVar()
resultStr.set(" ")

n1Label = Label (text="FFCS TIMETABLE MAKER").grid(row=1, column=3)

n1Label = Label (text="Enter the location of excel file").grid(row=2, column=1 ,sticky='wn')
n1Box = Entry()
n1Box.grid(row=2, column=2)
but1 = Button(text="Upload", command=excel).grid(row=2, column=3)


#//////////////////////////////Subject list				Subject list					Subject list////////////////////////////////////
	
global rp
rp=3
global cp
cp=1

n1Label = Label (text="Select the subjects : ").grid(row=rp, column=cp, sticky='wn')
lbs1 = Listbox(top,height=1)
cp+=1
lbs1.grid(row=rp,column=cp)
lbs1.insert(END, "Subject 1")


lbs2 = Listbox(top,height=1)
cp+=1
lbs2.grid(row=rp,column=cp)
lbs2.insert(END, "Subject 2")

lbs3 = Listbox(top,height=1)
cp+=1
lbs3.grid(row=rp,column=cp)
lbs3.insert(END, "Subject 3")

global fcp
fcp=cp

cp+=1
but2 = Button(text="OK", command=sel)
but2.grid(row=rp, column=cp)
cp+=1

global adds
adds=[]
adds.append(lbs1)
adds.append(lbs2)
adds.append(lbs3)

but3 = Button(text="More", command=add1)
but3.grid(row=rp, column=cp)







#////////////////////////////////////////////////////////      PREFERENCES   ////////////////////////////////////////////////////////////





n1Label = Label (text="Choose your preference if any:").grid(row=4, column=1,sticky='wn')


#VENUE
rp=5
cp5=0

C1 = Checkbutton(top, text = "Venue",  onvalue = 1, offvalue = 0, height=5, width = 20).grid(row=5, column=1,sticky=W)
lbv1 = Listbox(top,height=1)
lbv1.grid(row=5,column=2)
lbv1.insert(END, "GDN")
lbv1.insert(END, "MB")
lbv1.insert(END, "SJT")
lbv1.insert(END, "TT")
but = Button(text="More", command=excel).grid(row=5, column=3)


#TIME
rp=6
cp6=0

C2 = Checkbutton(top, text = "Time",  onvalue = 1, offvalue = 0, height=5, width = 20).grid(row=6, column=1,sticky=E)

t1Label = Label (text="From").grid(row=6, column=2)
lbt1 = Listbox(top,height=1)
for i in range(0,13):
	lbt1.insert(i, str(8+i) + ":00")
lbt1.grid(row=6,column=3,sticky=E)


t1Label = Label (text="To").grid(row=6, column=4)
lbt2 = Listbox(top,height=1)
for i in range(0,13):
	lbt2.insert(i, str(8+i) + ":00")	
lbt2.grid(row=6,column=5,sticky=E)
	
but = Button(text="More", command=excel)
but.grid(row=6, column=6,sticky=W)


#TEACHERS

rp=7
cp7=1
C3 = Checkbutton(top, text = "Teachers",  onvalue = 1, offvalue = 0,variable=mvar, height=5, width = 20).grid(row=7, column=1,sticky=W)

global lbst
lbst=[]

#add one row for st(subject-teacher preference)

lbst.append(Listbox(top,height=1))
lbst[0].grid(row=rp,column=2)
lbst[0].insert(1, "Select Subject")

but_ttupdate = Button(text="Update", command=ttupdate).grid(row=rp, column=3)

lbst.append(Listbox(top,height=1))
lbst[1].grid(row=rp,column=4)
lbst[1].insert(1, "Select Teacher")


global lbtt
lbtt=[]
lbtt.append(lbst)


global frp
frp=rp


#global but_tt
rp+=1
but_ttgen = Button(text="Generate timetable", command=generate)
but_ttgen.grid(row=rp, column=2)
	

lbtts=[]
but_addst = Button(text="More", command=addst)
but_addst.grid(row=7, column=5,sticky=E)



rp+=1
resultLabel = Label(textvariable=resultStr)
resultLabel.grid(row=rp, column=1)
resultStr.set("Error:\tNone")


top.mainloop()

